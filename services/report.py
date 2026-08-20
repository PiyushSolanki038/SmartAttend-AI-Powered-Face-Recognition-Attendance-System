from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import EXCEL_HEADERS
from db import queries

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

STATUS_FILL = {
    "present": GREEN_FILL,
    "absent": RED_FILL,
    "unknown": ORANGE_FILL,
    "manual": ORANGE_FILL,
}


def export_excel(filepath: str, subject: str = None, section: str = None, start_date: str = None, end_date: str = None,
                  department: str = None):
    rows = queries.get_filtered_attendance(subject, section, start_date, end_date, department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.merge_cells("A1:G1")
    ws["A1"] = "SmartAttend Attendance Report"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A2"] = f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["C2"] = f"Subject: {subject or 'All'}"
    ws["E2"] = f"Section: {section or 'All'}"

    header_row = 4
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    present_count = 0
    absent_count = 0
    total_count = 0

    row_idx = header_row + 1
    for row in rows:
        started = row["started_at"]
        date_part = started.strftime("%Y-%m-%d") if started else ""
        time_part = started.strftime("%H:%M:%S") if started else ""
        status = row["status"] or ""

        ws.cell(row=row_idx, column=1, value=row["roll_no"])
        ws.cell(row=row_idx, column=2, value=row["name"])
        ws.cell(row=row_idx, column=3, value=row["department"])
        ws.cell(row=row_idx, column=4, value=date_part)
        ws.cell(row=row_idx, column=5, value=time_part)
        ws.cell(row=row_idx, column=6, value=status)
        ws.cell(row=row_idx, column=7, value=row["confidence"])

        fill = STATUS_FILL.get(status)
        if fill:
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = fill

        total_count += 1
        if status == "present":
            present_count += 1
        elif status == "absent":
            absent_count += 1

        row_idx += 1

    attendance_pct = (present_count / total_count * 100) if total_count else 0
    summary_row = row_idx + 1
    ws.cell(row=summary_row, column=1,
            value=f"Total: {total_count} | Present: {present_count} | Absent: {absent_count} | Attendance %: {attendance_pct:.1f}%")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    for col_idx in range(1, 8):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    wb.save(filepath)
    return filepath
